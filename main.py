# main.py

import os
import openpyxl
from utils import get_user_input
from named_ranges import add_named_ranges
from formula_updater import update_formulas
import sys

def main():
    print("=== Excel Formula and Named Range Manager ===\n")
    
    # 1. Prompt the user for file path with default value
    default_file_path = '/Users/benjamin.van.dam/Documents/pit/Tax Calculation/Latest/Draft PB-berekening - FINAAL AJ24 - 23.09 - 22u49.xlsx'

    while True:
        file_path = get_user_input("Enter the full path to the Excel file", default_file_path)
        if not os.path.isfile(file_path):
            print(f"Error: The file '{file_path}' does not exist. Please enter a valid file path.\n")
            continue
        break

    # Load the workbook after validating the file path
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}\n")
        return

    # Main interaction loop
    while True:
        print("\n--- Main Menu ---")
        print("Please choose an option:")
        print("1. Create Named Ranges")
        print("2. Update Formulas")
        print("3. Save and Exit")
        print("4. Exit Without Saving")
        
        choice = get_user_input("Enter the number corresponding to your choice", "1")
        
        if choice == "1":
            # Create Named Ranges
            
            # Prompt for the sheet name to update named ranges
            default_specific_sheet = 'Tax Calculation'
            available_sheets = wb.sheetnames
            print(f"\nAvailable sheets: {', '.join(available_sheets)}")
            
            specific_sheet = get_user_input(
                f"Enter the sheet name to create/update named ranges (default: '{default_specific_sheet}')",
                default_specific_sheet
            ).strip()
            
            if specific_sheet not in wb.sheetnames:
                print(f"Error: Sheet '{specific_sheet}' does not exist in the workbook.")
                continue  # Return to the main menu if sheet doesn't exist
            
            ws = wb[specific_sheet]  # Set the worksheet for named range creation
            
            # Initialize a list to hold multiple configurations for named ranges
            configurations = []

            while True:
                print("\n--- Add a New Named Range Configuration ---")
                print("1. Named Range with Prefix")
                print("2. Named Range without Prefix")
                range_type = get_user_input("Enter the type of named range to create (1/2)", "1")

                if range_type == "1":
                    # Named Range with Prefix
                    default_prefix = 'display_code_'
                    default_cell_range = 'L200:L408'
                    default_search_columns = 'J,K'

                    prefix = get_user_input("Enter the prefix for the named ranges", default_prefix)
                    cell_range = get_user_input("Enter the cell range (e.g., 'L200:L408')", default_cell_range)
                    search_columns_input = get_user_input("Enter the columns to search for tax codes (e.g., 'J,K')", default_search_columns)
                    search_columns = [col.strip().upper() for col in search_columns_input.split(',')]

                    # Validate search columns
                    valid_columns = [chr(i) for i in range(ord('A'), ord('Z')+1)]
                    invalid_cols = [col for col in search_columns if col not in valid_columns]
                    if invalid_cols:
                        print(f"Error: Invalid column letters detected: {', '.join(invalid_cols)}. Please enter valid column letters (A-Z).\n")
                        continue

                    configurations.append({
                        'type': 'with_prefix',
                        'prefix': prefix,
                        'cell_range': cell_range,
                        'search_columns': search_columns
                    })
                elif range_type == "2":
                    # Named Range without Prefix
                    default_cell_range = 'L200:L408'
                    default_search_columns = 'J,K'

                    cell_range = get_user_input("Enter the cell range (e.g., 'L200:L408')", default_cell_range)
                    search_columns_input = get_user_input("Enter the columns to search for tax codes (e.g., 'J,K')", default_search_columns)
                    search_columns = [col.strip().upper() for col in search_columns_input.split(',')]

                    # Validate search columns
                    valid_columns = [chr(i) for i in range(ord('A'), ord('Z')+1)]
                    invalid_cols = [col for col in search_columns if col not in valid_columns]
                    if invalid_cols:
                        print(f"Error: Invalid column letters detected: {', '.join(invalid_cols)}. Please enter valid column letters (A-Z).\n")
                        continue

                    configurations.append({
                        'type': 'without_prefix',
                        'cell_range': cell_range,
                        'search_columns': search_columns
                    })
                else:
                    print("Invalid choice. Please enter 1 or 2.")
                    continue

                add_more = get_user_input("Do you want to add another range configuration? (yes/no)", "no").strip().lower()
                if add_more not in ['yes', 'y']:
                    break

            # Process each configuration after collecting all configurations
            for config in configurations:
                if config['type'] == 'with_prefix':
                    add_named_ranges(
                        wb=wb,
                        ws=ws,
                        prefix=config['prefix'],
                        cell_range=config['cell_range'],
                        search_columns=config['search_columns']
                    )
                else:
                    add_named_ranges(
                        wb=wb,
                        ws=ws,
                        cell_range=config['cell_range'],
                        search_columns=config['search_columns']
                    )

            print("\nNamed range creation completed.")
        
        elif choice == "2":
            # Update Formulas
            print("\n--- Updating Formulas ---")
            update_formulas(wb)
            print("Formula update completed.")
        
        elif choice == "3":
            # Save and Exit
            print("\n--- Saving Workbook ---")
            directory, file_name = os.path.split(file_path)
            
            # Prompt the user for saving preference
            save_choice = get_user_input(
                "Do you want to overwrite the original file or save as a new file? (overwrite/save_as_new)",
                "save_as_new"
            ).strip().lower()
            
            if save_choice in ['overwrite', 'o']:
                output_file = file_path  # Overwrite the original file
            else:
                new_file_name = 'updated_' + file_name
                output_file = os.path.join(directory, new_file_name)

            try:
                wb.save(output_file)
                if save_choice in ['overwrite', 'o']:
                    print(f"\nOriginal Excel file '{file_path}' has been overwritten.")
                else:
                    print(f"\nUpdated Excel file saved as '{output_file}'.")
            except Exception as e:
                print(f"Error saving workbook: {e}")
            print("Exiting the program. Goodbye!")
            break
        
        elif choice == "4":
            # Exit Without Saving
            confirm_exit = get_user_input("Are you sure you want to exit without saving? (yes/no)", "no").strip().lower()
            if confirm_exit in ['yes', 'y']:
                print("Exiting the program without saving. Goodbye!")
                break
            else:
                print("Returning to the main menu.")
        
        else:
            print("Invalid choice. Please enter a number between 1 and 4.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProgram interrupted. Exiting gracefully...")
        sys.exit(0)