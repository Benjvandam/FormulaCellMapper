import os
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

def get_user_input(prompt, default):
    """Helper function to get user input with a default value."""
    user_input = input(f"{prompt} (default: '{default}'): ")
    return user_input.strip() or default

def parse_cell(cell):
    """Parses a cell reference into column letter and row number."""
    col_letter = ''.join(filter(str.isalpha, cell)).upper()
    row_number = int(''.join(filter(str.isdigit, cell)))
    return col_letter, row_number

def add_named_ranges(wb, ws, prefix, cell_range, search_columns):
    """Adds named ranges based on the provided prefix, cell range, and search columns."""
    # Extract start and end cells
    try:
        start_cell, end_cell = cell_range.split(':')
    except ValueError:
        print(f"Error: Invalid cell range format '{cell_range}'. Please use format like 'L200:L408'.")
        return

    start_col_letter, start_row = parse_cell(start_cell)
    end_col_letter, end_row = parse_cell(end_cell)

    # Ensure target_column is correct (should always be the same as start_col_letter)
    target_column = start_col_letter

    # Debugging: Print parsed cell references
    print(f"\nProcessing Range:")
    print(f"Start Column: {start_col_letter}, Start Row: {start_row}")
    print(f"End Column: {end_col_letter}, End Row: {end_row}")
    print(f"Target Column: {target_column}")

    # Iterate over the range and find tax codes in the specified columns
    for row in range(start_row, end_row + 1):
        tax_code = None

        # Loop through each specified column to search for a tax code
        for col_letter in search_columns:
            cell_address = f'{col_letter}{row}'
            cell_value = ws[cell_address].value

            # Debugging: Uncomment the next line to see which cells are being checked
            print(f"Checking cell {cell_address} with value: {cell_value}")

            # Handle possible empty cells
            if cell_value is None:
                continue

            # Try converting to int (handles both int and float that can be cast to int)
            try:
                tax_code = int(cell_value)
                break  # Found a valid tax code
            except (ValueError, TypeError):
                continue  # Not a valid tax code, continue searching

        # If no valid tax code is found, skip this row
        if tax_code is None:
            continue

        # Check if the target cell has a valid value
        target_cell_address = f'{target_column}{row}'
        target_cell_value = ws[target_cell_address].value
        if target_cell_value is None:
            continue  # Skip if the target cell is empty

        # Create a named range for the target cell
        # Ensure sheet name is properly quoted (handles single quotes in sheet name)
        sheet_name_quoted = ws.title.replace("'", "''")
        target_cell_ref = f"'{sheet_name_quoted}'!${target_column}${row}"
        named_range = f"{prefix}{tax_code}"

        # Debugging: Print the named range details
        print(f"Creating named range '{named_range}' referring to '{target_cell_ref}'.")

        # Remove existing named range if it exists
        if named_range in wb.defined_names:
            del wb.defined_names[named_range]
            print(f"Removed existing named range '{named_range}'.")

        # Add the named range using DefinedName and append to defined_names
        try:
            # Ensure the formula starts with '='
            new_defined_name = DefinedName(name=named_range, attr_text=target_cell_ref)
            wb.defined_names[named_range] = new_defined_name  # Assign directly
            print(f"Named range '{named_range}' created successfully.")
        except Exception as e:
            print(f"Error creating named range '{named_range}': {e}")
            continue

def main():
    # 1. Prompt the user for file path and sheet name with default values
    default_file_path = '/Users/benjamin.van.dam/Documents/pit/Tax Calculation/Draft PB-berekening - AJ24 - Silverfin 31.07.2024.xlsx'
    default_sheet_name = 'Tax Calculation'

    file_path = get_user_input("Enter the full path to the Excel file", default_file_path)
    sheet_name = get_user_input("Enter the sheet name", default_sheet_name)

    # Load the Excel file and sheet
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' does not exist.")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
        return

    ws = wb[sheet_name]

    # List to hold multiple configurations
    configurations = []

    while True:
        print("\n--- Add a New Named Range Configuration ---")
        # 2. Prompt for other inputs with default values
        default_prefix = 'display_code_'
        default_cell_range = 'L200:L408'
        default_search_columns = 'J,K'

        prefix = get_user_input("Enter the prefix for the named cells", default_prefix)
        cell_range = get_user_input("Enter the cell range (e.g., 'L200:L408')", default_cell_range)
        search_columns_input = get_user_input("Enter the columns to search for tax codes (e.g., 'J,K')", default_search_columns)
        search_columns = [col.strip().upper() for col in search_columns_input.split(',')]

        # Validate search columns
        valid_columns = [chr(i) for i in range(ord('A'), ord('Z')+1)]
        invalid_cols = [col for col in search_columns if col not in valid_columns]
        if invalid_cols:
            print(f"Error: Invalid column letters detected: {', '.join(invalid_cols)}. Please enter valid column letters (A-Z).")
            continue

        # Store the configuration
        configurations.append({
            'prefix': prefix,
            'cell_range': cell_range,
            'search_columns': search_columns
        })

        # Ask the user if they want to add another configuration
        add_more = input("Do you want to add another range configuration? (yes/no): ").strip().lower()
        if add_more not in ['yes', 'y']:
            break

    # Process each configuration
    for config in configurations:
        add_named_ranges(
            wb=wb,
            ws=ws,
            prefix=config['prefix'],
            cell_range=config['cell_range'],
            search_columns=config['search_columns']
        )

    # 6. Save the updated workbook
    directory, file_name = os.path.split(file_path)
    new_file_name = 'updated_' + file_name
    output_file = os.path.join(directory, new_file_name)

    try:
        wb.save(output_file)
        print(f"\nUpdated Excel file saved as '{output_file}'.")
    except Exception as e:
        print(f"Error saving workbook: {e}")

if __name__ == "__main__":
    main()
