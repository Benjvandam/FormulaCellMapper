# formula_updater.py

import re
import time
from openpyxl.workbook.defined_name import DefinedName
from utils import get_user_input  # Assuming utils.py is in the same directory
from tqdm import tqdm  # Importing tqdm for progress indicators

def update_formulas(wb):
    """
    Updates formulas in selected worksheets by replacing cell references with their named ranges.
    Handles references with and without sheet names.

    For example:
    - In 'Tax Calculation' sheet: '=L404' becomes '=display_code_1657'
    - In 'Summary' sheet: '="Tax Calculation"!L404' becomes '=display_code_1657'

    Parameters:
    - wb: openpyxl Workbook object
    """
    # 1. Prompt the user to choose between updating a specific sheet or all sheets
    print("\n--- Formula Update Options ---")
    print("1. Update formulas in a specific sheet")
    print("2. Update formulas in all sheets")
    
    while True:
        choice = get_user_input("Enter the number corresponding to your choice (default: '1')", "1").strip()
        if choice not in ['1', '2']:
            print("Invalid choice. Please enter '1' or '2'.\n")
            continue
        break

    if choice == "1":
        # Update a specific sheet
        default_specific_sheet = 'Tax Calculation'
        available_sheets = wb.sheetnames
        print(f"\nAvailable sheets: {', '.join(available_sheets)}")
        specific_sheet = get_user_input(
            f"Enter the sheet name to update (default: '{default_specific_sheet}')",
            default_specific_sheet
        ).strip()
        if specific_sheet not in wb.sheetnames:
            print(f"Error: Sheet '{specific_sheet}' does not exist in the workbook.")
            return
        sheets_to_update = [wb[specific_sheet]]
    else:
        # Update all sheets
        sheets_to_update = wb.worksheets

    # 2. Create a nested mapping from sheet_name to (cell_address -> named_range)
    mapping = {}

    # Iterate over defined names
    for name in wb.defined_names:
        dn = wb.defined_names[name]  # Retrieve the DefinedName object
        if not isinstance(dn, DefinedName):
            print(f"Warning: '{name}' is not a DefinedName object.")
            continue
        print(f"Defined Name: '{dn.name}'")

        # Only consider named ranges that refer to single cells
        try:
            destinations = list(dn.destinations)  # List of (sheet, cell) tuples
        except AttributeError:
            print(f"Warning: DefinedName '{dn.name}' does not have 'destinations' attribute.")
            continue

        for sheet, coord in destinations:
            sheet_clean = sheet.strip("'").strip()
            coord_clean = coord.upper().replace('$', '')
            
            # Initialize the inner dictionary if the sheet is not yet in mapping
            if sheet_clean not in mapping:
                mapping[sheet_clean] = {}
            
            mapping[sheet_clean][coord_clean] = dn.name
            print(f"  -> Refers to: {sheet_clean}!{coord_clean}")

    # Debugging: Print the mapping
    print("\nMapping of Sheet and Cell Addresses to Named Ranges:")
    for sheet, cells in mapping.items():
        for cell, name in cells.items():
            print(f"{sheet}!{cell} => {name}")

    # Define regex to find cell references with optional sheet names
    # Matches patterns like:
    # 'Sheet1'!A1, Sheet1!A1, A1
    cell_ref_pattern = re.compile(r"(?:'([^']+)')?!?(\$?[A-Z]{1,3}\$?\d{1,7})")

    # 3. Define the replacement function
    def replace_match(match, current_sheet):
        sheet_name, cell_ref = match.groups()
        if sheet_name:
            # Reference includes sheet name
            ref_sheet = sheet_name
        else:
            # Reference does not include sheet name; assume current sheet
            ref_sheet = current_sheet
        cell_ref_clean = cell_ref.upper().replace('$', '')
        # Access the nested mapping
        return mapping.get(ref_sheet, {}).get(cell_ref_clean, match.group(0))  # Replace if mapped, else keep original

    # 4. Iterate through the selected sheets and update formulas
    for ws in tqdm(sheets_to_update, desc="Processing Sheets", unit="sheet"):
        # Check if the sheet is hidden
        if ws.sheet_state in ['hidden', 'veryHidden']:
            print(f"\nSkipping hidden sheet: {ws.title}")
            continue  # Skip processing this sheet

        print(f"\nProcessing sheet: {ws.title} at {time.strftime('%X')}")

        # **Step 1: Identify and Skip Non-Primary Merged Cells**
        # Create a set of all non-primary merged cells in the sheet
        non_primary_merged_cells = set()
        for merged_range in ws.merged_cells.ranges:
            # merged_range.bounds returns (min_col, min_row, max_col, max_row)
            min_col, min_row, max_col, max_row = merged_range.bounds
            # Iterate over all cells in the merged range except the primary cell
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_coord = ws.cell(row=row, column=col).coordinate
                    # Skip the primary cell (top-left cell)
                    if row == min_row and col == min_col:
                        continue
                    non_primary_merged_cells.add(cell_coord)

        # **Step 2: Preprocess Image Anchors to Create a Set of Cells with Images**
        image_cells = set()
        for image in ws._images:
            if hasattr(image.anchor, 'from_'):  # Note: In newer versions, it's 'from_'
                img_row = image.anchor.from_.row + 1  # zero-based index
                img_col = image.anchor.from_.col + 1  # zero-based index
                img_cell = ws.cell(row=img_row, column=img_col).coordinate
                image_cells.add(img_cell)

        # **Step 3: Calculate Total Number of Cells with Formulas to Process**
        # Exclude non-primary merged cells and cells with images from the total count
        total_formula_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.coordinate not in non_primary_merged_cells and cell.coordinate not in image_cells:
                    total_formula_cells += 1

        # **Step 4: Initialize Cell-Level Progress Bar**
        with tqdm(total=total_formula_cells, desc="Updating Cells", unit="cell", leave=False) as cell_pbar:
            for row in ws.iter_rows():
                for cell in row:
                    # **Step 4.1: Skip Non-Primary Merged Cells**
                    if cell.coordinate in non_primary_merged_cells:
                        continue  # Skip processing this cell

                    # **Step 4.2: Skip Cells Associated with Images**
                    if cell.coordinate in image_cells:
                        continue  # Skip processing this cell

                    if cell.data_type == 'f':  # 'f' indicates a formula
                        formula = cell.value
                        if not formula:
                            cell_pbar.update(1)
                            continue

                        # Check if the formula is a string
                        if not isinstance(formula, str):
                            print(f"Skipping cell {cell.coordinate} in '{ws.title}': Expected string formula, got {type(formula).__name__}")
                            cell_pbar.update(1)
                            continue

                        # Perform the replacement using re.sub with the replacement function
                        formula_new = cell_ref_pattern.sub(lambda m: replace_match(m, ws.title), formula)

                        if formula_new != formula:
                            # Inform the user about the formula update
                            print(f"  Updated cell {cell.coordinate} in '{ws.title}': '{formula}' to '{formula_new}'")
                            cell.value = formula_new

                        # Update the cell progress bar
                        cell_pbar.update(1)
